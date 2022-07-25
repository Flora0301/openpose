import openpyxl

wb = openpyxl.load_workbook('./key_frame2.xlsx')

# 获取所有的工作表
sheets = wb.sheetnames
print(sheets)

# 获取指定的工作表
sheet = sheets[0]
# 激活工作表
sheet = wb.active

cell = sheet.cell(1, 1)
cell.value = 'test'

# 保存
wb.save('./key_frame2.xlsx')