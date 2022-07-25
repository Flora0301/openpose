#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 28 19:17:47 2020

@author: joe
"""


from src import torch_openpose,util
import cv2
import openpyxl
import os


# video_path = './video/20220628140.mkv'
# cap = cv2.VideoCapture(video_path)

# wb = openpyxl.load_workbook('./key_frame2.xlsx')
# sheets = wb.sheetnames
# sheet = sheets[0]
# sheet = wb.active

for num in range(17, 18):
    # wb = openpyxl.load_workbook('./key_frame2.xlsx')
    # sheets = wb.sheetnames
    # sheet = sheets[0]
    # sheet = wb.active

    video = './video/20220628{0:03d}.mkv'.format(num)
    cap = cv2.VideoCapture(video)
    print("num: ", num)
    RBigToe_list = []
    img_num = 0

    while cap.isOpened():
        success, img = cap.read()
        img_num = img_num + 1
        if not success:
            print("Ignoring empty camera frame.")
            break
        img_resize = img[78:580, 280:1149]
        img_mul = img_resize.copy()

        tp = torch_openpose.torch_openpose('body_25')  # 载入模型
        poses, heels, heel, position = tp(img_resize)  # poses, heels 全部人的关键点，heel，position 目标人物关键点
        # print("position: " + str(position))

        # 脚趾关键点坐标
        RBigToe = position[0][22]
        print("RBigToe: " + str(RBigToe))
        RBigToe_list.append(RBigToe)

        img = util.draw_bodypose(img_resize, position, 'body_25')
        img_mul = util.draw_bodypose(img_mul, poses, 'body_25')
        cv2.imshow('v', img)

        filenum = num
        # 存储单人的关键点图片
        file = './pos/{0:03d}/'.format(filenum)
        fileName = './pos/{0:03d}/'.format(filenum) + str(img_num) + '.jpg'  # 存储路径
        # fileName_simple = './pos_image/{0:03d}/'.format(filenum) + str(num) + '(1)' + '.jpg'  # 存储路径
        # print("filename = " + fileName)
        # 判断路径是否存在，若不存在则创建
        if not os.path.exists(file):
            os.makedirs(file)
        cv2.imwrite(fileName, img, [cv2.IMWRITE_JPEG_QUALITY, 100])
        # cv2.imwrite(fileName_simple, image, [cv2.IMWRITE_JPEG_QUALITY, 100])

        # 存储多人的关键点图片
        file_mul = './pos_mul/{0:03d}/'.format(filenum)
        fileName_mul = './pos_mul/{0:03d}/'.format(filenum) + str(img_num) + '.jpg'  # 存储路径
        # fileName_simple = './pos_image/{0:03d}/'.format(filenum) + str(num) + '(1)' + '.jpg'  # 存储路径
        # print("filename = " + fileName)
        # 判断路径是否存在，若不存在则创建
        if not os.path.exists(file_mul):
            os.makedirs(file_mul)
        cv2.imwrite(fileName_mul, img_mul, [cv2.IMWRITE_JPEG_QUALITY, 100])

        if cv2.waitKey(5) & 0xFF == 27:
          break

    cap.release()
    cv2.destroyAllWindows()

    # 提取起始帧
    # print("list: " + str(RBigToe_list))

    num1 = len(RBigToe_list)
    dist = 0
    dist_list = []
    print('-----------------------------------{}'.format(len(dist_list)))
    frame_up = 0
    frame_down = 0
    next = 0

    x1 = 0
    y1 = 0
    # 计算距离
    for i in range(num1):
        x = RBigToe_list[i][0]
        y = RBigToe_list[i][1]
        dist = pow(pow(x-x1, 2) + pow(y-y1, 2), 0.5)
        dist_list.append(dist)
        x1 = x
        y1 = y
        print(str(i) + " dist: " + str(dist))

    print("dis_list: " + str(dist_list))
    for i in range(1, num1):
        if dist_list[i] > 11:
            frame_up = i
            print("起始帧： " + str(frame_up))
            break

    for j in range(frame_up, num1):
        if dist_list[j] < 10:
            frame_down = j + 1
            print("结束帧： " + str(frame_down))
            break

    # cell = sheet.cell(num+1, 2)
    # cell.value = str(frame_up)
    # cell1 = sheet.cell(num+1, 3)
    # cell1.value = str(frame_down)
    # wb.save('./key_frame2.xlsx')

# wb.save('./key_frame2.xlsx')
# for i in range(num):
#     if i % 2 == 0:
#         prior = RBigToe_list[i][1]
#         dist = next - prior
#         next = prior
#         if frame_up != 0 :
#             if dist > 10:
#                 frame_up = i
#
# # 提取结束帧
# for j in range(frame_up, num, 2):
#



# position: [[[297.0, 174.0, 1.0210559368133545], [267.0, 206.0, 0.9706293344497681], [241.0, 209.0, 0.9450704455375671],
#             [237.0, 276.0, 0.953989565372467], [251.0, 333.0, 0.9816102981567383], [293.0, 203.0, 0.9242585301399231],
#             [313.0, 259.0, 0.445705771446228], [0.0, 0.0, 0.0], [299.0, 325.0, 0.9471896886825562],
#             [280.0, 327.0, 0.9307724833488464], [286.0, 408.0, 0.8837556838989258], [294.0, 484.0, 0.9267120361328125],
#             [320.0, 323.0, 0.8936924934387207], [320.0, 405.0, 0.9090144634246826], [311.0, 477.0, 0.8894822001457214],
#             [290.0, 168.0, 0.9767276048660278], [299.0, 167.0, 0.6939061284065247], [263.0, 168.0, 1.000135898590088],
#             [0.0, 0.0, 0.0], [340.0, 493.0, 0.7072316408157349], [343.0, 486.0, 0.7370491623878479],
#             [306.0, 485.0, 0.5802323222160339], [325.0, 498.0, 0.8199432492256165], [307.0, 503.0, 0.7995603084564209],
#             [290.0, 492.0, 0.7928311228752136]]]


